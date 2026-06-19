import openpyxl
import json
import os
import sys
from datetime import datetime

def parse_score(score_str):
    if not score_str:
        return None
    score_str = str(score_str).strip()
    if '-' not in score_str:
        return None
    try:
        parts = score_str.split('-')
        return int(parts[0]), int(parts[1])
    except Exception:
        return None

def calculate_match_points(pred_str, act_str):
    pred = parse_score(pred_str)
    act = parse_score(act_str)
    
    if pred is None or act is None:
        return 0
        
    p_home, p_away = pred
    a_home, a_away = act
    
    # Rule 1: Exact correct prediction -> 45 points
    if p_home == a_home and p_away == a_away:
        return 45
        
    # Rule 2: Correct winner or draw (not exact) -> 30 points
    pred_winner = 1 if p_home > p_away else (-1 if p_home < p_away else 0)
    act_winner = 1 if a_home > a_away else (-1 if a_home < a_away else 0)
    
    if pred_winner == act_winner:
        return 30
        
    return 0

def clean_name(name):
    if not name:
        return ""
    name_str = str(name).strip()
    import re
    if re.search(r'Mbapp', name_str, re.IGNORECASE):
        return "K. Mbappé"
    return name_str

def main():
    # Use the directory containing the script to dynamically locate project files
    workspace = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(workspace, "voorspellingen.xlsx")
    results_path = os.path.join(workspace, "actual_results.json")
    leaderboard_path = os.path.join(workspace, "leaderboard.json")
    
    if not os.path.exists(xlsx_path):
        print(f"Error: Excel file not found at {xlsx_path}", file=sys.stderr)
        sys.exit(1)
        
    # Load actual results
    actual_results = {"matches": {}, "topscorers": {}, "champion": ""}
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                actual_results = json.load(f)
        except Exception as e:
            print(f"Warning: Could not parse actual_results.json: {e}. Using empty defaults.", file=sys.stderr)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except PermissionError:
        import subprocess
        temp_xlsx_path = xlsx_path + ".temp.xlsx"
        subprocess.run(["powershell", "-Command", f"Copy-Item '{xlsx_path}' '{temp_xlsx_path}'"], check=True)
        wb = openpyxl.load_workbook(temp_xlsx_path)
        try:
            os.remove(temp_xlsx_path)
        except Exception:
            pass
    sheet = wb.active
    
    # 1. Read participants (from row 1, col 2 onwards)
    participants = []
    max_c = sheet.max_column
    for col in range(2, max_c + 1):
        name = sheet.cell(row=1, column=col).value
        if name:
            participants.append({
                "name": name.strip(),
                "col_idx": col,
                "matches": {},
                "topscorers": [],
                "champion": ""
            })
            
    print(f"Found participants: {[p['name'] for p in participants]}")
    
    # 2. Read Match Predictions (Rows 2 to 73)
    matches_list = []
    for r in range(2, 74):
        match_id = sheet.cell(row=r, column=1).value
        if not match_id:
            continue
        match_id = match_id.strip()
        matches_list.append(match_id)
        
        # Get actual score
        act_score = actual_results.get("matches", {}).get(match_id, "")
        
        for p in participants:
            pred_score = sheet.cell(row=r, column=p["col_idx"]).value
            if pred_score is not None:
                pred_score = str(pred_score).strip()
            else:
                pred_score = ""
            pts = calculate_match_points(pred_score, act_score)
            p["matches"][match_id] = {
                "prediction": pred_score,
                "points": pts
            }

    # 3. Read Topscorer Predictions (Rows 74 to 79)
    topscorer_keys = ["Topscorer1", "Topscorer2", "Topscorer3", "Topscorer4", "Topscorer5", "Topscorer6"]
    for idx, r in enumerate(range(74, 80)):
        key = sheet.cell(row=r, column=1).value or topscorer_keys[idx]
        for p in participants:
            player_name = sheet.cell(row=r, column=p["col_idx"]).value
            player_name = clean_name(player_name)
            p["topscorers"].append(player_name)

    # 4. Read Champion Prediction (Row 80)
    for p in participants:
        p["champion"] = clean_name(sheet.cell(row=80, column=p["col_idx"]).value)

    # Auto-register new topscorers from predictions to actual_results.json if not present
    actual_results_changed = False
    if "topscorers" not in actual_results:
        actual_results["topscorers"] = {}
    
    for p in participants:
        for ts_name in p["topscorers"]:
            if not ts_name:
                continue
            cleaned_pred_name = clean_name(ts_name)
            found = False
            for k in actual_results["topscorers"].keys():
                if clean_name(k).lower() == cleaned_pred_name.lower():
                    found = True
                    break
            if not found:
                actual_results["topscorers"][cleaned_pred_name] = {
                    "goals": 0,
                    "position": "attacker"
                }
                actual_results_changed = True
                print(f"Registered new predicted topscorer in actual_results: {cleaned_pred_name}")
                
    if actual_results_changed:
        try:
            with open(results_path, "w", encoding="utf-8") as f:
                json.dump(actual_results, f, indent=2, ensure_ascii=False)
            print(f"Updated actual_results.json with new topscorer(s) at: {results_path}")
        except Exception as e:
            print(f"Error saving actual_results.json: {e}", file=sys.stderr)

    # 5. Calculate Points
    # Topscorer points multiplier mapping
    multiplier_map = {
        "attacker": 8,
        "midfielder": 16,
        "defender": 32
    }
    
    # Actual topscorers goals and positions
    actual_topscorers = actual_results.get("topscorers", {})
    actual_champion = clean_name(actual_results.get("champion", ""))
    
    # Build list of unique topscorers for the summary
    unique_topscorers_summary = {}
    for name, data in actual_topscorers.items():
        cleaned_n = clean_name(name)
        pos = data.get("position", "attacker").lower()
        goals = int(data.get("goals", 0))
        mult = multiplier_map.get(pos, 8)
        unique_topscorers_summary[cleaned_n] = {
            "name": cleaned_n,
            "position": pos,
            "goals": goals,
            "pointsPerGoal": mult,
            "totalPoints": goals * mult
        }

    leaderboard_records = []
    
    for p in participants:
        # Sum match points
        match_pts = sum(m["points"] for m in p["matches"].values())
        
        # Calculate topscorer points
        ts_details = []
        ts_pts = 0
        total_picked_goals = 0
        for ts_name in p["topscorers"]:
            # Find in actual results
            # Try to match case-insensitively or with normalized encoding
            actual_data = None
            matched_key = ts_name
            for k in actual_topscorers.keys():
                if clean_name(k).lower() == ts_name.lower():
                    actual_data = actual_topscorers[k]
                    matched_key = clean_name(k)
                    break
            
            if actual_data:
                goals = int(actual_data.get("goals", 0))
                pos = actual_data.get("position", "attacker").lower()
                mult = multiplier_map.get(pos, 8)
                player_pts = goals * mult
                total_picked_goals += goals
            else:
                goals = 0
                pos = "attacker"
                mult = 8
                player_pts = 0
                
            ts_pts += player_pts
            ts_details.append({
                "name": matched_key,
                "position": pos,
                "goals": goals,
                "pointsPerGoal": mult,
                "points": player_pts
            })
            
        # Calculate champion points
        correct_champion = False
        champ_pts = 0
        if actual_champion and p["champion"].lower() == actual_champion.lower():
            correct_champion = True
            champ_pts = 250
            
        total_pts = match_pts + ts_pts + champ_pts
        
        leaderboard_records.append({
            "name": p["name"],
            "matchPoints": match_pts,
            "topscorerPoints": ts_pts,
            "championPoints": champ_pts,
            "totalPoints": total_pts,
            "totalGoals": total_picked_goals,
            "predictions": {
                "matches": p["matches"],
                "topscorers": ts_details,
                "champion": {
                    "predicted": p["champion"],
                    "actual": actual_champion,
                    "points": champ_pts,
                    "correct": correct_champion
                }
            }
        })

    # Sort leaderboard records
    leaderboard_records.sort(key=lambda x: x["totalPoints"], reverse=True)
    # Add ranks
    for rank, record in enumerate(leaderboard_records, 1):
        record["rank"] = rank

    # Format full matches list
    matches_comparison = []
    for match_id in matches_list:
        act_score = actual_results.get("matches", {}).get(match_id, "")
        match_pred_details = {}
        for p in participants:
            match_pred_details[p["name"]] = {
                "prediction": p["matches"][match_id]["prediction"],
                "points": p["matches"][match_id]["points"]
            }
        matches_comparison.append({
            "id": match_id,
            "actual": act_score,
            "predictions": match_pred_details
        })

    # Compile the final leaderboard JSON
    output_data = {
        "lastUpdated": datetime.now().isoformat(),
        "leaderboard": [{
            "rank": r["rank"],
            "name": r["name"],
            "matchPoints": r["matchPoints"],
            "topscorerPoints": r["topscorerPoints"],
            "championPoints": r["championPoints"],
            "totalPoints": r["totalPoints"],
            "totalGoals": r["totalGoals"]
        } for r in leaderboard_records],
        "participantsDetails": leaderboard_records,
        "matches": matches_comparison,
        "topscorers": list(unique_topscorers_summary.values()),
        "champion": {
            "actual": actual_champion
        }
    }
    
    with open(leaderboard_path, "w", encoding="utf-8") as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
        
    print(f"Successfully calculated standings and wrote leaderboard.json at: {leaderboard_path}")

if __name__ == "__main__":
    main()
